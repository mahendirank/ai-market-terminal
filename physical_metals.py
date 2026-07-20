"""
physical_metals.py — ETF vault-flow tracker for gold & silver (supply-chain Phase 1).

Free primary sources (probed live 2026-07-20 before integration):
- GLD tonnes-in-trust, full daily history:
    https://api.spdrgoldshares.com/api/v1/historical-archive?product=gld&exchange=NYSE&lang=en  (XLSX)
- SLV tonnes-in-trust, current snapshot only (iShares exposes no archive endpoint):
    scraped from the embedded JSON on the product page — daily snapshots
    accumulate here in sqlite, so SLV history builds forward from install date.

Why this matters for trading: daily ETF tonnage flows are the best free proxy
for institutional *physical* accumulation vs distribution. The 20-day net flow
direction, set against price, flags divergences (price down + tonnes up =
accumulation into weakness).
"""
import os, re, html, sqlite3, threading, time
from datetime import datetime, timezone

import requests

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "db", "physical_cache.db")
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/126.0 Safari/537.36"}
GLD_URL = "https://api.spdrgoldshares.com/api/v1/historical-archive?product=gld&exchange=NYSE&lang=en"
SLV_URL = "https://www.ishares.com/us/products/239855/ishares-silver-trust-fund"

REFRESH_SECS = 4 * 3600          # sources update once per trading day
_lock = threading.Lock()
_last_refresh = 0.0


def _conn():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("""CREATE TABLE IF NOT EXISTS holdings (
        metal  TEXT NOT NULL,
        d      TEXT NOT NULL,          -- ISO date
        tonnes REAL NOT NULL,
        PRIMARY KEY (metal, d))""")
    return conn


def _fetch_gld_rows():
    """Parse the SPDR historical-archive XLSX → [(iso_date, tonnes), ...]."""
    import io
    import openpyxl
    resp = requests.get(GLD_URL, timeout=30, headers=HEADERS)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
    # First sheet is a disclaimer; the data sheet is named like "US GLD Historical Archive"
    ws = next((wb[s] for s in wb.sheetnames if "archive" in s.lower()), wb[wb.sheetnames[-1]])
    rows = ws.iter_rows(values_only=True)
    date_i = tonnes_i = None
    out = []
    for row in rows:
        if date_i is None:
            cells = [str(c or "").lower() for c in row]
            for i, c in enumerate(cells):
                if date_i is None and "date" in c:
                    date_i = i
                if tonnes_i is None and "tonne" in c:
                    tonnes_i = i
            continue
        try:
            d, t = row[date_i], row[tonnes_i]
            if d is None or t in (None, ""):
                continue
            if isinstance(d, datetime):
                iso = d.date().isoformat()
            else:
                iso = datetime.strptime(str(d).strip(), "%d-%b-%Y").date().isoformat()
            out.append((iso, float(str(t).replace(",", ""))))
        except (ValueError, TypeError, IndexError):
            continue
    return out


def _fetch_slv_snapshot():
    """Scrape (iso_date, tonnes) from the embedded JSON on the SLV page."""
    resp = requests.get(SLV_URL, timeout=30, headers=HEADERS)
    resp.raise_for_status()
    text = html.unescape(resp.text)
    m = re.search(r'"tonnes":\{[^}]*?"formattedValue":"([\d,\.]+)"', text)
    if not m:
        raise ValueError("tonnes not found on SLV page")
    tonnes = float(m.group(1).replace(",", ""))
    md = re.search(r'"formattedAsOfDate":"([A-Za-z]{3} \d{1,2}, \d{4})","name":"tonnes"', text)
    if md:
        iso = datetime.strptime(md.group(1), "%b %d, %Y").date().isoformat()
    else:
        iso = datetime.now(timezone.utc).date().isoformat()
    return iso, tonnes


def _refresh_if_due(force=False):
    global _last_refresh
    with _lock:
        if not force and time.time() - _last_refresh < REFRESH_SECS:
            return
        _last_refresh = time.time()   # set first: a failing source retries next window, not every call
    conn = _conn()
    try:
        gld = _fetch_gld_rows()[-120:]
        conn.executemany("INSERT OR REPLACE INTO holdings VALUES ('GOLD', ?, ?)", gld)
        conn.commit()
    except Exception as e:
        print(f"[physical] GLD refresh failed: {type(e).__name__}: {e}", flush=True)
    try:
        iso, tonnes = _fetch_slv_snapshot()
        conn.execute("INSERT OR REPLACE INTO holdings VALUES ('SILVER', ?, ?)", (iso, tonnes))
        conn.commit()
    except Exception as e:
        print(f"[physical] SLV refresh failed: {type(e).__name__}: {e}", flush=True)
    conn.close()


def _series(conn, metal, n=30):
    rows = conn.execute(
        "SELECT d, tonnes FROM holdings WHERE metal=? ORDER BY d DESC LIMIT ?",
        (metal, n)).fetchall()
    return list(reversed(rows))


def _metal_block(conn, metal, etf):
    s = _series(conn, metal)
    if not s:
        return {"etf": etf, "status": "no-data"}
    latest_d, latest_t = s[-1]
    chg_1d = round(latest_t - s[-2][1], 2) if len(s) >= 2 else None
    flow_5d = round(latest_t - s[-6][1], 2) if len(s) >= 6 else None
    flow_20d = round(latest_t - s[-21][1], 2) if len(s) >= 21 else None
    if flow_20d is None:
        label = "HISTORY BUILDING"
    elif flow_20d > latest_t * 0.001:
        label = "ACCUMULATION"
    elif flow_20d < -latest_t * 0.001:
        label = "DISTRIBUTION"
    else:
        label = "FLAT"
    return {
        "etf": etf, "status": "ok",
        "tonnes": round(latest_t, 2), "as_of": latest_d,
        "chg_1d": chg_1d, "flow_5d": flow_5d, "flow_20d": flow_20d,
        "label": label, "points": len(s),
        "series": [round(t, 2) for _, t in s],
    }


def get_physical_metals() -> dict:
    """Vault-flow snapshot for the dashboard PHYSICAL panel."""
    _refresh_if_due()
    conn = _conn()
    try:
        return {
            "gold":   _metal_block(conn, "GOLD", "GLD"),
            "silver": _metal_block(conn, "SILVER", "SLV"),
            "updated": datetime.now(timezone.utc).isoformat(),
        }
    finally:
        conn.close()


if __name__ == "__main__":
    _refresh_if_due(force=True)
    import json
    print(json.dumps(get_physical_metals(), indent=2))
